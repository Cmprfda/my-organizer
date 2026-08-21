# -*- coding: utf-8 -*-
"""Camada de dados/serviço: constrói o que a interface consome."""

import hashlib
import json
import os
import re
import time
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from . import config
from . import events
from .config import (APP_VERSION, BASE_STATUSES, CANDIDATE_DIRS, DEFAULT_PERSON,
                     DEFAULT_SHEET, HERE, lan_ip)
from .statefile import read_json, write_json
from .excel import (_ADMIN_CACHE, _RAW_CACHE, admin_statuses, close_excel_workbook,
                    detect_header_row, find_named_file, find_tracker_files,
                    forget_files_cache, load_sheet_snapshot, locate_row_in,
                    pick_sheet, set_data_validation_fixed_list,
                    set_data_validation_list, write_cells_to_excel)
from .graph import (GRAPH_PATH, GraphError, current_book, graph_config, graph_forget_item,
                    graph_ids_from_path, graph_load_rows, graph_modified, graph_path_for,
                    graph_state, graph_state_public, has_book, is_graph_path)
from .history import HISTORY_COLS, mark_app_write, record_read
from .i18n import msg
from .logs import log_event
from .team import load_team_waiting
from .store import (load_ccrs, load_notes, load_overrides, load_waiting,
                    save_notes, save_overrides)
from .text import cell_to_text, normalize, person_matcher
from .todos import load_todo, save_todo

# última leitura bem-sucedida por (ficheiro, aba, pessoa, todas) — serve de
# fallback quando o Excel tem o ficheiro bloqueado em exclusivo
_LAST_GOOD = {}

# --------------------------------------------------------------------------
# O retrato da última leitura que correu bem, gravado no disco
#
# O `_LAST_GOOD` acima vive só em MEMÓRIA: servia o livro bloqueado pelo Excel
# durante a sessão, mas arrancar a app sem rede (no trem, com o OneDrive em
# baixo, com o VPN fora) dava uma vista de Tarefas vazia — e vazio parece "não
# tens nada", não "não consegui ler". Toda a história de trabalhar offline já
# existia (as alterações ✎ ficam locais, a lista Por fazer e as notas são
# locais); o que faltava era sobreviver ao reinício do processo.
#
# Não é uma cache: é um retrato de leitura, e a app di-lo em cima da vista com a
# hora a que foi tirado. Nunca é escrito no Excel nem conta como valor da folha.
# --------------------------------------------------------------------------

LAST_READ_FILE = os.path.join(HERE, "last_read.json")
# retratos guardados (livro+aba+pessoa+vista): mais do que isto é peso no disco
# sem valor — ninguém arranca às escuras para ver a sexta combinação
LAST_READ_KEEP = 6
# um retrato maior do que isto não se guarda: uma folha enorme faria o arranque
# (que é quando isto é lido) esperar pelo disco em vez de mostrar a app
LAST_READ_MAX_BYTES = 3 * 1024 * 1024
# campos que não fazem sentido num retrato: são sobre o estado desta sessão e
# são recalculados a cada leitura
LAST_READ_SKIP = ("files", "graph", "cell_view", "notice", "warning", "error",
                  "hint", "filter_lists")


def _snapshot_key(cache_key):
    """A chave do _LAST_GOOD como texto, para poder ir para um JSON."""
    return "||".join(str(p) for p in cache_key)


# as impressões digitais do que está gravado no disco, semeadas de uma vez por
# ficheiro. O caso normal — leitura igual à anterior, de dois em dois minutos por
# cada janela aberta — sai por aqui sem reler o last_read.json inteiro (que
# chega perto de 100 KB) só para comparar oito carateres.
#
# O mapa é validado contra a marca do ficheiro (existência, hora e tamanho), e
# não só contra o nome: apagar o last_read.json à mão é um passo de recuperação
# que a app documenta, e um mapa a dizer "isso já está gravado" sobre um
# ficheiro que já não existe deixava o arranque sem rede sem retrato até o
# conteúdo da folha mudar. Um os.stat por pedido é o que isto custa — a leitura
# do ficheiro inteiro é que era o peso.
_LAST_READ_DIGESTS = (None, None, {})   # (ficheiro, marca, mapa)


def _last_read_mark():
    try:
        st = os.stat(LAST_READ_FILE)
        return (st.st_mtime_ns, st.st_size)
    except OSError:
        return None      # não existe (ou não se alcança): não há nada gravado


def _last_read_digests():
    global _LAST_READ_DIGESTS
    marca = _last_read_mark()
    if _LAST_READ_DIGESTS[0] != LAST_READ_FILE or _LAST_READ_DIGESTS[1] != marca:
        semente = {k: v.get("digest") for k, v in _read_snapshots().items()
                   if isinstance(v, dict) and v.get("digest")} if marca else {}
        _LAST_READ_DIGESTS = (LAST_READ_FILE, marca, semente)
    return _LAST_READ_DIGESTS[2]


def _remember_last_read_mark():
    """Depois de gravar, a marca do mapa passa a ser a do ficheiro novo."""
    global _LAST_READ_DIGESTS
    _LAST_READ_DIGESTS = (LAST_READ_FILE, _last_read_mark(), _LAST_READ_DIGESTS[2])


def save_last_read(cache_key, result):
    """Grava o retrato desta leitura (só se ele mudou desde o último)."""
    if not isinstance(result, dict) or result.get("error") or not result.get("rows"):
        return
    chave = _snapshot_key(cache_key)
    digest = result.get("digest")
    digests = _last_read_digests()
    # a impressão digital do conteúdo já é calculada para a vista: com ela, uma
    # leitura igual à anterior (o caso normal, a cada 2 minutos) não escreve nada
    if digest and digests.get(chave) == digest:
        return
    guardados = _read_snapshots()
    retrato = {k: v for k, v in result.items() if k not in LAST_READ_SKIP}
    retrato["at"] = datetime.now().replace(microsecond=0).isoformat()
    try:
        tamanho = len(json.dumps(retrato, ensure_ascii=False).encode("utf-8"))
    except (TypeError, ValueError):
        return
    if tamanho > LAST_READ_MAX_BYTES:
        log_event(f"retrato de {chave[:60]} grande demais ({tamanho // 1024} KB) - não guardado")
        return
    guardados[chave] = retrato
    # os mais recentes ficam: a ordem é a da hora a que cada um foi tirado
    if len(guardados) > LAST_READ_KEEP:
        ordem = sorted(guardados, key=lambda k: str(guardados[k].get("at") or ""))
        for velho in ordem[:len(guardados) - LAST_READ_KEEP]:
            guardados.pop(velho, None)
            digests.pop(velho, None)
    write_json(LAST_READ_FILE, guardados)
    if digest:
        digests[chave] = digest
    else:
        digests.pop(chave, None)
    _remember_last_read_mark()


def _read_snapshots():
    data = read_json(LAST_READ_FILE, {})
    return data if isinstance(data, dict) else {}


def load_last_read(cache_key):
    """O retrato guardado desta combinação, ou None."""
    retrato = _read_snapshots().get(_snapshot_key(cache_key))
    if not isinstance(retrato, dict) or not retrato.get("rows"):
        return None
    return retrato


def forget_last_read():
    """Esquece os retratos gravados (linha de comandos e testes)."""
    global _LAST_READ_DIGESTS
    write_json(LAST_READ_FILE, {})
    _LAST_READ_DIGESTS = (None, None, {})


# marca de versão da fonte na altura em que cada entrada do _RAW_CACHE foi
# enchida, pela mesma chave (ficheiro, aba). Enquanto a marca não mudar, as
# linhas em cache SÃO as linhas do livro: reler o .xlsx inteiro (ou voltar a
# pedir todas as linhas à nuvem) daria exatamente o mesmo.
_RAW_STAMP = {}

# colunas fixas do tracker, pelo nome canónico (ver col_by_name em read_sheet):
# o valor atual destas vai em row_meta["cur"], para o cartão do TODO mostrar a
# linha ao vivo mesmo com a folha fora da vista
TRACKER_COLS = ("Function/TC", "To Do", "OBS", "Status TC", "Status TP")


def forget_cache(path=None):
    """Esquece o que foi lido (de um ficheiro, ou de todos): a leitura seguinte
    vai buscar tudo de novo, como se a app tivesse acabado de abrir."""
    for cache in (_RAW_CACHE, _RAW_STAMP, _LAST_GOOD):
        for key in [k for k in cache if k and (path is None or k[0] == path)]:
            cache.pop(key, None)
    for key in [k for k in _SYNC_CHECK if path is None or k[0] == path]:
        _SYNC_CHECK.pop(key, None)
    if path is None:
        _ADMIN_CACHE.clear()
        forget_last_read()
        # "Atualizar" relê tudo de raiz: também as pastas onde os livros são
        # procurados, senão um ficheiro acabado de lá pôr só aparecia a seguir
        forget_files_cache()
    else:
        _ADMIN_CACHE.pop(path, None)


def forget_web_cache():
    """Esquece tudo o que foi lido da fonte web (usado ao trocar de livro no
    OneDrive: os dados em cache eram do livro anterior). Apanha todos os livros
    lidos nesta sessão, não só o GRAPH_PATH sem livro indicado."""
    paths = {GRAPH_PATH}
    for cache in (_RAW_CACHE, _LAST_GOOD):
        paths |= {k[0] for k in cache if k and is_graph_path(k[0])}
    paths |= {p for p in _ADMIN_CACHE if is_graph_path(p)}
    for path in paths:
        forget_cache(path)


def local_twin(files, book=None):
    """Cópia sincronizada do livro do OneDrive (o mesmo ficheiro numa pasta
    local do OneDrive), ou None. Vale a pena preferi-la: as alterações feitas
    no Excel aparecem no disco assim que são gravadas, enquanto a cópia na
    nuvem só as recebe quando o OneDrive acaba de sincronizar (pode demorar
    minutos). Sem `book` procura a cópia do livro escolhido na app."""
    name = ((book if book is not None else current_book()) or {}).get("name") or ""
    for p in files:
        if os.path.basename(p).lower() == name.lower():
            return p
    return None


def known_files(book=None):
    """Ficheiros que a app aceita abrir: os candidatos habituais mais a cópia
    sincronizada do livro do OneDrive (que pode ter outro nome). Sem `book` é o
    livro escolhido na app."""
    files = find_tracker_files()
    name = ((book if book is not None else current_book()) or {}).get("name") or ""
    if name and not any(os.path.basename(p).lower() == name.lower() for p in files):
        files += find_named_file(name)
        files.sort(key=os.path.getmtime, reverse=True)
    return files


def rows_digest(rows):
    """Impressão digital do conteúdo de uma folha crua (texto das células).
    Ignora células e linhas vazias no fim, porque o Excel local e a nuvem não
    contam da mesma maneira onde a folha acaba."""
    limpas = []
    for r in (rows or []):
        linha = [cell_to_text(c) for c in (r or [])]
        while linha and not linha[-1].strip():
            linha.pop()
        limpas.append(linha)
    while limpas and not limpas[-1]:
        limpas.pop()
    return hashlib.md5(json.dumps(limpas, ensure_ascii=False).encode("utf-8")).hexdigest()


# (ficheiro, aba) -> (marca da nuvem, mtime local, cópias diferentes)
_SYNC_CHECK = {}


def sync_gap(path, sheet, local_rows, mtime, drive_id="", item_id=""):
    """True quando a cópia na nuvem tem conteúdo diferente do ficheiro local,
    ou seja, o OneDrive ainda não acabou de sincronizar (numa direção ou na
    outra). As datas não servem para isto: o OneDrive atualiza a data do item
    na nuvem antes de o conteúdo novo lá estar. Só se compara de facto quando
    uma das cópias mudou, para não ler o livro da nuvem a cada pedido."""
    key = (path, normalize(sheet))
    _, tag = graph_modified(drive_id, item_id)
    antes = _SYNC_CHECK.get(key)
    if antes and antes[0] == tag and antes[1] == mtime:
        return antes[2]
    _, _, cloud_rows = graph_load_rows(drive_id, item_id, sheet)
    diferentes = rows_digest(cloud_rows) != rows_digest(local_rows)
    _SYNC_CHECK[key] = (tag, mtime, diferentes)
    if diferentes:
        log_event(f"a cópia no OneDrive de {os.path.basename(path)} ainda difere "
                  f"da local (sincronização a decorrer)")
    return diferentes


def _wb_key(workbook_id, sheet, fn, todo):
    """Identidade de uma linha nos overrides/notas: livro||aba||função||to do.
    O livro faz parte da chave para que a mesma linha em livros diferentes não
    partilhe estados nem notas."""
    return f"{workbook_id}||{sheet}||{fn}||{todo}"


def _legacy_key(sheet, fn, todo):
    """Formato antigo (aba||função||to do), de quando só havia um livro."""
    return f"{sheet}||{fn}||{todo}"


def _override_entry(overrides, workbook_id, sheet, fn, todo):
    """(chave, valor) desta linha em `overrides`/`notes`: a chave com livro ou,
    se ainda não existir, a antiga sem livro — os ficheiros gravados antes
    desta versão continuam a valer, sem serem reescritos."""
    key = _wb_key(workbook_id, sheet, fn, todo)
    if key in overrides:
        return key, overrides[key]
    antiga = _legacy_key(sheet, fn, todo)
    if antiga in overrides:
        return antiga, overrides[antiga]
    return key, None


def _split_key(key):
    """(livro, aba, função, to do) de uma chave de override. Nas chaves antigas
    (três partes) não há livro: fica None."""
    parts = str(key).split("||")
    if len(parts) >= 4:
        return parts[0], parts[1], parts[2], "||".join(parts[3:])
    parts += [""] * (3 - len(parts))
    return None, parts[0], parts[1], "||".join(parts[2:])


def _source_stamp(path):
    """Marca de versão da fonte (muda a cada gravação do livro), sem ler a
    folha. None quando não se consegue saber — nesse caso relê-se sempre.

    Não é a marca que vai para a interface (essa é a do current_stamp, ao
    segundo, para o cliente comparar): esta só se compara consigo própria, por
    isso leva a data ao nanossegundo e o tamanho — duas gravações no mesmo
    segundo dariam a mesma marca ao segundo e a segunda passaria despercebida.
    """
    try:
        if is_graph_path(path):
            return graph_modified(*graph_ids_from_path(path))[1] or None
        st = os.stat(path)
        return f"{st.st_mtime_ns}|{st.st_size}"
    except Exception:
        return None


def read_sheet(path, sheet_name, person, show_all, lang="pt"):
    raw_key = (path, normalize(sheet_name))
    warning_ts = None
    # a folha crua só se relê quando o livro foi gravado desde a última leitura.
    # Sem isto, cada /api/tasks reabria o .xlsx inteiro (ou voltava a pedir
    # todas as linhas à nuvem) para obter exatamente as mesmas linhas — e a
    # interface pede-o de dois em dois minutos por cada livro aberto. Poupa-se
    # só a leitura CRUA: tudo o que vem a seguir (overrides, notas, "à espera
    # de alguém", histórico) continua a correr a cada pedido, porque muda sem o
    # livro mudar.
    stamp = _source_stamp(path)
    em_cache = _RAW_CACHE.get(raw_key)
    raw_hit = bool(em_cache and stamp and _RAW_STAMP.get(raw_key) == stamp)
    if raw_hit:
        _, real_sheet, all_sheets, rows = em_cache
    else:
        try:
            if is_graph_path(path):
                drive_id, item_id = graph_ids_from_path(path)
                real_sheet, all_sheets, rows = graph_load_rows(drive_id, item_id, sheet_name)
                if real_sheet is None:
                    return {"error": msg("err_nosheet", lang, s=sheet_name),
                            "sheets": all_sheets}
            else:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                try:
                    all_sheets = wb.sheetnames
                    real_sheet = pick_sheet(wb, sheet_name)
                    if real_sheet is None:
                        return {"error": msg("err_nosheet", lang, s=sheet_name),
                                "sheets": all_sheets}
                    ws = wb[real_sheet]
                    rows = [list(r) for r in ws.iter_rows(values_only=True)]
                finally:
                    wb.close()
            _RAW_CACHE[raw_key] = (datetime.now(), real_sheet, all_sheets, rows)
            # sem marca (não se conseguiu saber a versão da fonte) não se guarda
            # nada: a leitura seguinte volta a ler, como antes desta cache
            if stamp:
                _RAW_STAMP[raw_key] = stamp
            else:
                _RAW_STAMP.pop(raw_key, None)
        except Exception as exc:
            # ficheiro bloqueado pelo Excel (ou a meio da sincronização):
            # continua com a última leitura crua — os filtros e as edições de
            # estado aplicam-se na mesma
            if raw_key not in _RAW_CACHE:
                raise
            # a leitura falhou: a marca guardada já não descreve o que está em
            # cache, senão o pedido seguinte servia isto como se fosse fresco
            _RAW_STAMP.pop(raw_key, None)
            warning_ts, real_sheet, all_sheets, rows = _RAW_CACHE[raw_key]
            log_event(f"leitura de {path} falhou ({exc!r}) - "
                      f"a servir a leitura das {warning_ts:%H:%M}")

    header_index = detect_header_row(rows)
    if header_index is None:
        return {"error": msg("err_noheader", lang, s=real_sheet),
                "sheets": all_sheets}

    raw_headers = rows[header_index]
    headers = [cell_to_text(h) or f"Coluna {i + 1}" for i, h in enumerate(raw_headers)]

    # índices das colunas usadas para chaves e para os overrides (estados e OBS)
    hidx = {}
    for j, h in enumerate(headers):
        hn = normalize(h)
        if hn in ("function/tc", "to do", "status tc", "status tp", "obs"):
            hidx[hn] = j
    for j, h in enumerate(headers):
        hn = normalize(h)
        if hn in ("author tc", "reviewer tc", "author tp", "reviewer tp"):
            hidx[hn] = j
    # nome real (verbatim) de cada coluna -> índice — para a vista mapeada à
    # medida (viewmap.js), onde a coluna editada é identificada pelo próprio
    # texto do cabeçalho, não por um nome fixo do tracker. Os nomes fixos têm
    # prioridade quando coincidem com o texto real de outra coluna.
    col_by_name = {h: j for j, h in enumerate(headers)}
    for want, col_name in (("status tc", "Status TC"), ("status tp", "Status TP"),
                           ("obs", "OBS"), ("function/tc", "Function/TC"), ("to do", "To Do")):
        if want in hidx:
            col_by_name[col_name] = hidx[want]
    overrides = load_overrides()
    notes = load_notes()
    waiting_all = load_waiting()
    # esperas que os colegas publicaram na pasta partilhada (ver team.py): a
    # minha marca manda sempre na linha; onde eu não tenho nenhuma, vale a de
    # quem já está a cobrar aquilo — é o que evita duas pessoas a cobrar a mesma
    # linha e a mesma pessoa
    waiting_team = load_team_waiting(person)

    # folha sem as colunas do tracker (qualquer outro livro de Excel): mostra-se
    # tal como está, sem filtrar por pessoa nem sincronizar papéis
    generic = "function/tc" not in hidx
    if generic:
        show_all = True

    person_norm = normalize(person) if person else ""
    # aceita também células só com um dos nomes (ex.: "Carlos"), porque a folha
    # usa nomes inconsistentes ("Mariana" vs "Mariana Ribeiro"). O teste vive no
    # text.py para as esperas da equipa usarem o MESMO (ver team.team_waiting_on)
    mentions_person = person_matcher(person)

    def is_me(cells, idx_name):
        if idx_name not in hidx:
            return False
        return mentions_person(cells[hidx[idx_name]])

    def applicable(status_text):
        t = normalize(status_text)
        return t not in ("", "n/a")

    data_rows, row_meta, statuses = [], [], set()
    # retrato desta leitura para o histórico: TODAS as linhas da folha, não só as
    # que passam o filtro da pessoa (ver record_read, cswaios/history.py)
    history_rows = []
    total_rows = 0
    overrides_stale = False
    for i, row in enumerate(rows[header_index + 1:]):
        xlrow = header_index + 2 + i  # nº da linha real na folha (1-based)
        cells = [cell_to_text(v) for v in row]
        # ignora linhas completamente vazias
        if not any(c for c in cells):
            continue
        total_rows += 1
        cells += [""] * (len(headers) - len(cells))

        fn_key = cells[hidx["function/tc"]] if "function/tc" in hidx else ""
        todo_key = cells[hidx["to do"]] if "to do" in hidx else ""
        # valor cru de qualquer coluna real desta linha — serve de "base" a
        # qualquer alteração local, seja de uma coluna fixa do tracker (Status
        # TC/TP, OBS, Function/TC, To Do) seja de uma coluna mapeada na vista
        # personalizada (viewmap.js), identificada pelo seu próprio texto
        orig = {name: (cells[j] if j < len(cells) else "") for name, j in col_by_name.items()}
        over = {}
        if orig.get("Status TC"):
            statuses.add(orig["Status TC"])
        if orig.get("Status TP"):
            statuses.add(orig["Status TP"])
        # o histórico segue sempre o valor CRU da folha (orig), nunca o `cells`
        # com a alteração local aplicada: senão editar um estado aqui aparecia
        # duas vezes, uma ao editar e outra quando o Push a leva à folha
        history_rows.append({"xlrow": xlrow, "fn": fn_key, "todo": todo_key,
                             "cols": {c: orig.get(c, "") for c in HISTORY_COLS}})

        # papel por vertente (usado para sincronizar TODO por regras de autoria/review)
        role_sync = {"author": [], "reviewer": []}
        st_tc = cells[hidx["status tc"]] if "status tc" in hidx else ""
        st_tp = cells[hidx["status tp"]] if "status tp" in hidx else ""
        if applicable(st_tc):
            if is_me(cells, "author tc"):
                role_sync["author"].append(st_tc)
            if is_me(cells, "reviewer tc"):
                role_sync["reviewer"].append(st_tc)
        if applicable(st_tp):
            if is_me(cells, "author tp"):
                role_sync["author"].append(st_tp)
            if is_me(cells, "reviewer tp"):
                role_sync["reviewer"].append(st_tp)
        okey, entry = _override_entry(overrides, path, real_sheet, fn_key, todo_key)
        if entry:
            for col_name, o in list(entry.items()):
                j = col_by_name.get(col_name)
                if j is None or not isinstance(o, dict):
                    continue
                if cells[j] == o.get("base", ""):
                    cells[j] = str(o.get("value", ""))
                    over[col_name] = True
                else:
                    # a folha mudou desde o override: a folha ganha, e o override
                    # é removido de vez (senão ressuscitava se a célula voltasse
                    # ao valor antigo)
                    entry.pop(col_name)
                    overrides_stale = True
            if not entry:
                overrides.pop(okey)

        # valor ATUAL das colunas fixas do tracker (já com qualquer alteração
        # local aplicada, ao contrário do `orig`, que guarda sempre o valor cru
        # da folha para o Push poder comparar): é daqui que o cartão do TODO lê
        # o título/"o que fazer"/estados ao vivo, sem precisar da folha à vista
        # (ver liveTaskContent, static/js/todo.js)
        cur = {name: (cells[j] if j < len(cells) else "")
               for name, j in col_by_name.items() if name in TRACKER_COLS}

        # quem está ligado à linha (autor/reviewer de cada vertente): texto cru
        # da folha, só para mostrar — não entra nos overrides nem na escrita
        people = {}
        for want, key in (("author tc", "author_tc"), ("reviewer tc", "reviewer_tc"),
                          ("author tp", "author_tp"), ("reviewer tp", "reviewer_tp")):
            people[key] = cells[hidx[want]].strip() if want in hidx else ""

        if show_all or not person_norm or any(mentions_person(c) for c in cells if c):
            _, note = _override_entry(notes, path, real_sheet, fn_key, todo_key)
            # "à espera de alguém" desta linha (ver load_waiting): vai com a
            # linha para o cliente poder mostrar o chip e não a contar como
            # parada enquanto a espera for razoável
            _, waiting = _override_entry(waiting_all, path, real_sheet, fn_key, todo_key)
            if not isinstance(waiting, dict):
                # a chave partilhada não leva o livro: o caminho do ficheiro é
                # diferente em cada máquina (ver _shared_key, team.py)
                waiting = waiting_team.get(_legacy_key(real_sheet, fn_key, todo_key))
            data_rows.append(cells[:len(headers)])
            row_meta.append({"fn": fn_key, "todo": todo_key, "orig": orig, "over": over,
                             "cur": cur, "note": note, "xlrow": xlrow,
                             "people": people,
                             "waiting": waiting if isinstance(waiting, dict) else None,
                             "todo_sync_role": role_sync})

    if overrides_stale:
        save_overrides(overrides)

    # remove colunas sem qualquer valor nas linhas apresentadas
    if data_rows:
        keep = [i for i in range(len(headers))
                if headers[i].strip() and any(i < len(r) and r[i] for r in data_rows)]
        headers = [headers[i] for i in keep]
        data_rows = [[r[i] if i < len(r) else "" for i in keep] for r in data_rows]

    # dropdown de estados: a lista oficial da aba Admin, mais quaisquer valores
    # extra que existam nas células (ex.: N/A), no fim
    oficial = admin_statuses(path) or BASE_STATUSES
    status_list = oficial + sorted(statuses - set(oficial), key=str.lower)

    # colunas reais na folha (1-based), para a escrita via Excel/COM
    xlcols = {}
    for want, name in (("status tc", "Status TC"), ("status tp", "Status TP"),
                       ("obs", "OBS"), ("function/tc", "fn"),
                       ("function/tc", "Function/TC"), ("to do", "To Do")):
        if want in hidx:
            xlcols[name] = hidx[want] + 1

    # histórico: só com uma leitura FRESCA (com warning_ts estamos a servir o
    # retrato antigo do _RAW_CACHE — as linhas são as mesmas de propósito) e só
    # numa folha do tracker (numa folha genérica nenhuma destas colunas existe).
    # Nunca deixar o histórico rebentar uma leitura: é um extra, não o serviço.
    #
    # Com a marca da fonte igual à da cache (raw_hit) as linhas são as MESMAS
    # que a leitura fresca que pôs a marca já anotou: voltar a abrir e comparar
    # o history.json inteiro a cada pedido só para concluir "não mudou nada"
    # custava o ficheiro todo por cada janela aberta, de dois em dois minutos.
    # Depois de um Push o `forget_cache` limpa a marca, por isso a leitura
    # seguinte é fresca e a escrita continua a ser atribuída à app.
    if warning_ts is None and not raw_hit and not generic:
        try:
            record_read(path, real_sheet, history_rows)
        except Exception as exc:
            log_event(f"não consegui anotar o histórico de {real_sheet} ({exc!r})")

    warn_key = "warning_web" if is_graph_path(path) else "warning_locked"
    return {
        "warning": msg(warn_key, lang, t=f"{warning_ts:%H:%M}") if warning_ts else None,
        "sheet": real_sheet,
        "sheets": all_sheets,
        "xlcols": xlcols,
        "headers": headers,
        "rows": data_rows,
        "row_meta": row_meta,
        "statuses": status_list,
        "total_rows": total_rows,
        "person": person,
    }


_CELL_REF_RE = re.compile(r"^([A-Za-z]{1,3})(\d+)$")


def parse_cell_ref(ref):
    """Converte uma referência estilo Excel ("C3") em (linha0, coluna0), ambos
    0-based. None se o formato for inválido."""
    m = _CELL_REF_RE.match(str(ref or "").strip())
    if not m:
        return None
    try:
        col0 = column_index_from_string(m.group(1).upper()) - 1
    except ValueError:
        return None
    return int(m.group(2)) - 1, col0


def _ensure_raw_rows(path, sheet_wanted):
    """Garante que a folha pedida está no _RAW_CACHE (lendo-a se preciso) e
    devolve (aba real, grelha em bruto) — usado para ler a lista de opções de
    uma categoria (ver _read_list_options), que pode viver numa aba diferente
    da que está a ser vista."""
    raw_key = (path, normalize(sheet_wanted))
    cached = _RAW_CACHE.get(raw_key)
    if cached:
        return cached[1], cached[3]
    try:
        if is_graph_path(path):
            drive_id, item_id = graph_ids_from_path(path)
            real_sheet, all_sheets, rows = graph_load_rows(drive_id, item_id, sheet_wanted)
            if real_sheet is None:
                return None, None
        else:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                real_sheet = pick_sheet(wb, sheet_wanted)
                if real_sheet is None:
                    return None, None
                rows = [list(r) for r in wb[real_sheet].iter_rows(values_only=True)]
                all_sheets = wb.sheetnames
            finally:
                wb.close()
    except Exception:
        return None, None
    _RAW_CACHE[raw_key] = (datetime.now(), real_sheet, all_sheets, rows)
    return real_sheet, rows


def _read_list_options(path, sheet_wanted, cell, orientation, size):
    """Valores de uma lista de opções predefinida (aba + célula inicial +
    orientação + tamanho), para o dropdown de uma categoria com useList=true
    na vista mapeada à medida (ver build_cell_categories)."""
    ref = parse_cell_ref(cell)
    if ref is None or not sheet_wanted:
        return []
    row0, col0 = ref
    try:
        size = max(1, int(size))
    except (TypeError, ValueError):
        size = 1
    real_sheet, rows = _ensure_raw_rows(path, sheet_wanted)
    if real_sheet is None:
        return []

    def cell_at(r0, c0):
        if 0 <= r0 < len(rows):
            row = rows[r0]
            if 0 <= c0 < len(row):
                return cell_to_text(row[c0])
        return ""

    if orientation == "vertical":
        vals = [cell_at(row0 + k, col0) for k in range(size)]
    else:
        vals = [cell_at(row0, col0 + k) for k in range(size)]
    seen, out = set(), []
    for v in vals:
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _cellcat_key(workbook_id, sheet, xlrow, col0):
    """Identidade de uma célula de categoria livre (vista mapeada à medida):
    livro||aba||__cellcat__||linha||coluna. Ao contrário das colunas fixas do
    tracker (identificadas por Function/TC+To Do), uma folha genérica pode não
    ter nenhuma das duas, por isso usa-se a própria posição na folha — tal
    como metaByRow, no cliente, já usa o xlrow em vez de função+"to do" por
    ser a única coisa realmente única por linha."""
    return f"{workbook_id}||{sheet}||__cellcat__||{xlrow}||{col0}"


def _split_cellcat_key(key):
    """Inverso de _cellcat_key: (livro, aba, linha, coluna), ou None se a
    chave não seguir esse formato (ver push_overrides)."""
    parts = str(key).split("||")
    if len(parts) != 5 or parts[2] != "__cellcat__":
        return None, None, None, None
    return parts[0], parts[1], parts[3], parts[4]


def queue_cellcat_override(workbook_id, sheet, xlrow, col0, value, base, list_cfg):
    """Alteração local (✎) a uma categoria livre com lista predefinida (vista
    mapeada à medida, ver build_cell_categories): fica em `overrides.json`
    até ao Push, tal como as colunas fixas do tracker. `list_cfg` (aba+célula+
    orientação+tamanho da lista de opções) viaja com o próprio override para
    o Push também poder aplicar a validação nativa do Excel a essa célula
    (ver push_overrides), sem precisar de reconsultar a configuração da vista."""
    overrides = load_overrides()
    key = _cellcat_key(workbook_id, sheet, xlrow, col0)
    if value is None:
        overrides.pop(key, None)
    else:
        entry = {"value": str(value), "base": base}
        if list_cfg:
            entry["list"] = list_cfg
        overrides[key] = entry
    save_overrides(overrides)


def build_cell_categories(path, sheet_name, categories, row_meta):
    """Vista resumida à medida por coordenadas de célula (ver viewmap.js): cada
    categoria lê a célula em startCell (ou, com `size` > 1, essa célula mais as
    seguintes na mesma linha se orientation="horizontal", ou na mesma coluna se
    "vertical") e concatena-as. `size` vazio/nulo lê só a própria célula — uma
    categoria == uma célula do Excel, sem tentar adivinhar onde a próxima
    categoria começa. Lê sempre da grelha em bruto desta leitura (a mesma que
    alimentou read_sheet), nunca de data_rows/headers já truncados e podados
    (tasks.py:311/321-325), porque só a grelha em bruto preserva a posição real
    das colunas.

    Todas as categorias são editáveis (ver cellCatHtml/openCellCatEditor,
    tasks.js): sem useList, como texto livre, tal como a OBS ou o Function/TC.
    Uma categoria com useList=true fica antes limitada a uma lista de valores
    predefinida, vinda de uma de duas fontes: um intervalo do próprio livro
    (listSheet+listCell+listOrientation+listSize, listMode="range") ou uma
    lista fixa guardada na biblioteca por aba do cliente (listMode="fixed",
    valores já resolvidos em listValues, ver viewmap.js). A alteração fica
    local (✎) até ao Push, tal como as colunas fixas do tracker, mas guardada
    com uma chave própria (ver _cellcat_key) porque folhas genéricas não têm
    Function/TC nem To Do para identificar a linha."""
    cached = _RAW_CACHE.get((path, normalize(sheet_name)))
    if not cached or not categories:
        return {"headers": [], "rows": []}
    _, real_sheet, _, rows = cached

    def cell_at(r0, c0):
        if 0 <= r0 < len(rows):
            row = rows[r0]
            if 0 <= c0 < len(row):
                return cell_to_text(row[c0])
        return ""

    headers, specs = [], []
    for i, cat in enumerate(categories or []):
        if not isinstance(cat, dict):
            continue
        ref = parse_cell_ref(cat.get("startCell"))
        if ref is None:
            continue
        row0, col0 = ref
        orientation = "vertical" if cat.get("orientation") == "vertical" else "horizontal"
        raw_size = cat.get("size")
        size = 1
        if raw_size not in (None, ""):
            try:
                size = max(1, int(raw_size))
            except (TypeError, ValueError):
                size = 1
        name = str(cat.get("name") or "").strip() or cell_at(row0, col0) or f"Categoria {i + 1}"
        list_cfg = None
        if cat.get("useList") and cat.get("listMode") == "fixed":
            # biblioteca de listas predefinidas por aba (ver viewmap.js): o
            # cliente já resolveu listId -> valores literais antes de mandar
            # este cellcat, porque a biblioteca só existe no localStorage dele
            values = cat.get("listValues")
            if isinstance(values, list) and values:
                list_cfg = {"fixed": True, "values": [str(v) for v in values]}
        elif cat.get("useList") and cat.get("listSheet") and cat.get("listCell"):
            list_cfg = {
                "sheet": str(cat.get("listSheet") or ""),
                "cell": str(cat.get("listCell") or ""),
                "orientation": "horizontal" if cat.get("listOrientation") == "horizontal" else "vertical",
                "size": cat.get("listSize"),
            }
        headers.append(name)
        specs.append((col0, orientation, size, list_cfg))

    def _list_options(cfg):
        if not cfg:
            return []
        if cfg.get("fixed"):
            seen, out = set(), []
            for v in cfg.get("values") or []:
                if v and v not in seen:
                    seen.add(v)
                    out.append(v)
            return out
        return _read_list_options(path, cfg["sheet"], cfg["cell"], cfg["orientation"], cfg["size"])

    options = [_list_options(cfg) for _, _, _, cfg in specs]
    use_list_flags = [cfg is not None for _, _, _, cfg in specs]

    overrides = load_overrides()
    overrides_stale = False
    data_rows, pending_rows, base_rows = [], [], []
    for meta in row_meta or []:
        xlrow = meta["xlrow"]
        r0 = xlrow - 1
        line, pending_line, base_line = [], [], []
        for col0, orientation, size, list_cfg in specs:
            base = cell_at(r0, col0)
            # todas as categorias são editáveis, com ou sem lista predefinida
            # (ver cellCatHtml/openCellCatEditor, tasks.js): sem lista, a
            # alteração é texto livre, tal como a OBS ou o Function/TC
            key = _cellcat_key(path, real_sheet, xlrow, col0)
            entry = overrides.get(key)
            pending = False
            if entry and entry.get("base", "") == base:
                value, pending = str(entry.get("value", "")), True
            else:
                if entry:
                    # a folha mudou desde a alteração local: descarta-a, tal
                    # como as colunas fixas do tracker fazem em read_sheet
                    overrides.pop(key, None)
                    overrides_stale = True
                if size > 1:
                    vals = [cell_at(r0, col0 + k) for k in range(size)] if orientation == "horizontal" \
                        else [cell_at(r0 + k, col0) for k in range(size)]
                    value = " ".join(v for v in vals if v)
                else:
                    value = base
            line.append(value)
            pending_line.append(pending)
            base_line.append(base)
        data_rows.append(line)
        pending_rows.append(pending_line)
        base_rows.append(base_line)

    if overrides_stale:
        save_overrides(overrides)

    cols = [col0 for col0, _, _, _ in specs]
    lists = [list_cfg for _, _, _, list_cfg in specs]
    return {"headers": headers, "rows": data_rows, "useList": use_list_flags,
            "options": options, "pending": pending_rows, "base": base_rows,
            "cols": cols, "lists": lists}


def known_headers(path, sheet):
    """Cabeçalhos (texto verbatim, com o mesmo fallback "Coluna N" que read_sheet
    expõe ao cliente) da última leitura em cache desta folha, ou None se ainda
    não foi lida nesta sessão. Usado só para validar nomes de coluna vindos do
    cliente em /api/update, sem reler o livro."""
    want = normalize(sheet)
    cached = _RAW_CACHE.get((path, want))
    if not cached:
        # a chave do _RAW_CACHE é a aba *pedida* à read_sheet, que pode não ser
        # a que ficou resolvida (aba vazia = a habitual, ou a primeira do livro);
        # o cliente manda sempre a resolvida, por isso procura-se também por essa
        for (cpath, _), entry in _RAW_CACHE.items():
            if cpath == path and normalize(entry[1]) == want:
                cached = entry
                break
    if not cached:
        return None
    _, _, _, rows = cached
    header_index = detect_header_row(rows)
    if header_index is None:
        return None
    return [cell_to_text(h) or f"Coluna {i + 1}" for i, h in enumerate(rows[header_index])]


def _relink_row(workbook_id, sheet, fn, todo, new_fn, new_todo):
    """Depois de o Function/TC ou o "To Do" mudarem mesmo na folha, refaz as
    ligações que usam a identidade da linha (livro||aba||função||to do): a nota
    fixada na tarefa e os itens do TODO que apontam para ela. Sem isto, as
    ligações partiam-se em silêncio no Push."""
    new_key = _wb_key(workbook_id, sheet, new_fn, new_todo)

    notes = load_notes()
    # a nota pode ainda estar na chave antiga (sem livro), de antes desta versão
    old_key, _ = _override_entry(notes, workbook_id, sheet, fn, todo)
    if old_key in notes:
        notes[new_key] = notes.pop(old_key)
        save_notes(notes)

    todos = load_todo()
    changed = False
    sheet_norm = normalize(sheet)
    for item in todos:
        if not isinstance(item, dict):
            continue
        links = item.get("links") if isinstance(item.get("links"), list) else []
        refs = [item.get("ref")] + [lk.get("ref") for lk in links if isinstance(lk, dict)]
        for ref in refs:
            if not isinstance(ref, dict):
                continue
            # aba em branco = ligação antiga sem aba guardada; conta como
            # coincidência (mesma convenção do sync com o TODO em todos.py)
            ref_sheet = str(ref.get("sheet") or "")
            if ref_sheet and normalize(ref_sheet) != sheet_norm:
                continue
            if str(ref.get("fn") or "") != fn or str(ref.get("todo") or "") != todo:
                continue
            ref["fn"] = new_fn
            ref["todo"] = new_todo
            changed = True
    if changed:
        save_todo(todos)


def push_overrides(target):
    """Escreve no Excel/OneDrive as alterações de estado guardadas localmente.
    O livro de destino é sempre explícito — a app não escolhe nenhum por si.
    Devolve (ficheiro, enviadas, falhadas). Usado pelo /api/push e pela linha
    de comandos."""
    if not target:
        raise ValueError("ficheiro desconhecido")
    if is_graph_path(target):
        if not graph_config():
            raise ValueError("ficheiro desconhecido")
    elif os.path.normcase(target) not in {os.path.normcase(p) for p in known_files()}:
        raise ValueError("ficheiro desconhecido")
    overrides = load_overrides()
    pushed, failed = 0, []
    # etiqueta deste Push: viaja com cada célula escrita e volta nos eventos do
    # histórico, para o "desfazer" poder pegar no envio inteiro em vez de ir
    # célula a célula (ver mark_app_write e /api/history/undo)
    batch = f"p{int(time.time() * 1000)}"
    # chaves novas já usadas nesta chamada (linhas renomeadas): impede que duas
    # linhas diferentes, ambas renomeadas para a mesma identidade neste mesmo
    # Push, fundam por engano as suas colunas pendentes numa só
    renamed_this_call = set()

    # O Push faz-se em duas voltas. Na primeira junta-se tudo o que há para
    # escrever — uma leitura da folha por aba, em vez de uma releitura do livro
    # inteiro por cada alteração pendente — e na segunda arruma-se o resultado
    # (histórico, renomeações, o que falhou). Pelo meio há uma só ida ao Excel
    # por aba: antes era um PowerShell e uma gravação do livro POR CÉLULA, o que
    # punha um envio de cinco linhas a demorar dezenas de segundos.
    jobs = {}       # normalize(aba) -> (grafia da aba, [grupos de linha])
    plano = []      # o que fazer com cada chave, na ordem original
    preset = {}     # resultados já decididos sem ir ao Excel
    snapshots = {}  # normalize(aba) -> leitura da folha (ou None)
    # linhas que não estavam na folha lida: dizem-se no fim, quando já se sabe
    # se alguma renomeação deste envio explica a ausência
    nao_encontradas = []
    renomeadas = set()   # abas onde este envio mudou algum Function/TC
    seq = 0

    def _juntar(sheet, grupo):
        # a aba fica com a primeira grafia que aparecer: o Worksheets.Item do
        # Excel não distingue maiúsculas de minúsculas
        nsheet = normalize(sheet)
        if nsheet not in jobs:
            jobs[nsheet] = (sheet, [])
        jobs[nsheet][1].append(grupo)

    for key in list(overrides.keys()):
        entry = overrides.get(key)
        if not isinstance(entry, dict):
            continue
        if "||__cellcat__||" in key:
            # categoria livre com lista predefinida (vista mapeada à medida):
            # identificada por posição na folha, não por Function/TC+To Do
            # (ver _cellcat_key) — escreve-se diretamente, sem locate_row
            wb_id, sheet, xlrow_s, col0_s = _split_cellcat_key(key)
            if wb_id is None or os.path.normcase(wb_id) != os.path.normcase(target):
                continue
            try:
                xlrow, col0 = int(xlrow_s), int(col0_s)
            except (TypeError, ValueError):
                overrides.pop(key, None)
                continue
            xlcol = col0 + 1
            valor, base = entry.get("value", ""), entry.get("base", "")
            # a célula é a sua própria "guarda": só escreve se ainda tiver o
            # texto que foi lido quando a alteração local foi feita
            _juntar(sheet, {"xlrow": xlrow, "fncol": xlcol, "guard": base,
                            "cells": [{"i": seq, "xlcol": xlcol, "value": valor}]})
            plano.append({"kind": "cellcat", "key": key, "entry": entry, "sheet": sheet,
                          "xlrow": xlrow, "col0": col0, "xlcol": xlcol,
                          "valor": valor, "i": seq})
            seq += 1
            continue
        wb_id, sheet, fn, todo = _split_key(key)
        # chave antiga (sem livro): é do tempo em que só havia um, vai para o
        # destino pedido; com livro, só se for mesmo este
        if wb_id is not None and os.path.normcase(wb_id) != os.path.normcase(target):
            continue
        nsheet = normalize(sheet)
        if nsheet not in snapshots:
            snapshots[nsheet] = load_sheet_snapshot(target, sheet)
        snap = snapshots[nsheet]
        coords = locate_row_in(snap, fn, todo) if snap else None
        if coords is None:
            # A folha é lida uma vez, ANTES de se escrever: uma linha que só
            # passe a ter esta identidade por causa de uma renomeação deste
            # mesmo envio ainda não está aqui (antes, com uma releitura do livro
            # por cada alteração, aparecia). Não se perde nada — a alteração
            # fica pendente e o envio seguinte, já com a folha relida, leva-a —
            # mas quem está a olhar merece que se lhe diga isso.
            nao_encontradas.append({"fn": fn, "sheet": nsheet})
            continue
        xlrow, hidx = coords
        cells, cols = [], []
        for col_name, o in list(entry.items()):
            want = normalize(col_name)
            if want not in hidx or not isinstance(o, dict):
                # não chega a ir ao Excel, mas guarda-se o lugar na fila para a
                # falha sair na ordem certa (e com a guarda certa) na 2ª volta
                preset[seq] = (False, f"coluna {col_name} não encontrada")
                cols.append((seq, col_name, None))
            else:
                valor = o.get("value", "")
                cells.append({"i": seq, "xlcol": hidx[want] + 1, "value": valor})
                cols.append((seq, col_name, valor))
            seq += 1
        if cells:
            _juntar(sheet, {"xlrow": xlrow, "fncol": hidx["function/tc"] + 1,
                            "guard": fn, "cells": cells})
        plano.append({"kind": "row", "key": key, "entry": entry, "sheet": sheet,
                      "wb_id": wb_id, "fn": fn, "todo": todo, "xlrow": xlrow,
                      "cols": cols})

    # ---- a ida ao Excel: um PowerShell e uma gravação do livro por aba ----
    results = dict(preset)
    for grafia, grupos in jobs.values():
        results.update(write_cells_to_excel(target, grafia, grupos))

    # ---- 2ª volta: arrumar o resultado, na ordem original das chaves ----
    for rec in plano:
        key, entry, sheet = rec["key"], rec["entry"], rec["sheet"]
        if rec["kind"] == "cellcat":
            xlrow, col0, xlcol = rec["xlrow"], rec["col0"], rec["xlcol"]
            valor = rec["valor"]
            ok, msg_text = results.get(rec["i"], (False, "sem resposta do Excel"))
            if ok:
                overrides.pop(key, None)
                pushed += 1
                # para o histórico saber que esta mudança na folha foi daqui
                headers_now = known_headers(target, sheet)
                mark_app_write(target, sheet, xlrow,
                               headers_now[col0] if headers_now and col0 < len(headers_now)
                               else f"Coluna {xlcol}", valor, batch)
                list_cfg = entry.get("list")
                if isinstance(list_cfg, dict) and list_cfg.get("fixed") and list_cfg.get("values"):
                    set_data_validation_fixed_list(
                        target, sheet, f"{get_column_letter(xlcol)}{xlrow}", list_cfg.get("values"))
                elif isinstance(list_cfg, dict) and list_cfg.get("sheet") and list_cfg.get("cell"):
                    set_data_validation_list(
                        target, sheet, f"{get_column_letter(xlcol)}{xlrow}",
                        list_cfg.get("sheet"), list_cfg.get("cell"),
                        list_cfg.get("orientation"), list_cfg.get("size"))
            else:
                failed.append({"fn": f"{sheet}!{get_column_letter(xlcol)}{xlrow}", "error": msg_text})
            continue
        wb_id, fn, todo, xlrow = rec["wb_id"], rec["fn"], rec["todo"], rec["xlrow"]
        # o Function/TC e o "To Do" fazem parte da identidade da linha: se
        # forem escritos, a identidade muda e há ligações a refazer
        new_fn, new_todo = fn, todo
        # a escrita confirma sempre a célula do Function/TC antes de gravar;
        # depois de a mudarmos, as restantes colunas desta linha têm de ser
        # confirmadas com o valor novo (o script faz o mesmo, célula a célula)
        guard_fn = fn
        for i, col_name, valor in rec["cols"]:
            ok, msg_text = results.get(i, (False, "sem resposta do Excel"))
            if ok:
                entry.pop(col_name, None)
                pushed += 1
                mark_app_write(target, sheet, xlrow, col_name, valor, batch)
                if col_name == "Function/TC":
                    new_fn = guard_fn = str(valor)
                    renomeadas.add(normalize(sheet))
                elif col_name == "To Do":
                    new_todo = str(valor)
            else:
                failed.append({"fn": guard_fn, "error": msg_text})
        if (new_fn, new_todo) != (fn, todo):
            # a identidade nova passa a incluir o livro, mesmo que a antiga
            # ainda não o tivesse (chave anterior a esta versão)
            wb_key = wb_id if wb_id is not None else target
            if entry:
                # ainda sobram colunas por enviar nesta linha: passam para a
                # chave nova, senão ficavam órfãs (a leitura seguinte calcula a
                # identidade a partir do conteúdo novo da folha) — a não ser que
                # outra linha deste mesmo Push já tenha sido renomeada para a
                # mesma identidade nova, caso em que fundir destruiria as
                # colunas de uma das duas linhas: fica por enviar e falha, em
                # vez de contaminar a linha errada
                new_key = _wb_key(wb_key, sheet, new_fn, new_todo)
                if new_key in renamed_this_call:
                    for col_name in list(entry.keys()):
                        failed.append({"fn": guard_fn, "error":
                                       f"coluna {col_name} não enviada: outra linha "
                                       "deste Push ficou com a mesma identidade nova"})
                    entry = {}
                    overrides.pop(key, None)
                else:
                    overrides.pop(key, None)
                    destino = overrides.get(new_key)
                    if not isinstance(destino, dict):
                        destino = {}
                        overrides[new_key] = destino
                    destino.update(entry)
                    entry = destino
                    renamed_this_call.add(new_key)
            else:
                renamed_this_call.add(_wb_key(wb_key, sheet, new_fn, new_todo))
            _relink_row(wb_key, sheet, fn, todo, new_fn, new_todo)
        if not entry:
            overrides.pop(key, None)

    # as linhas que não se acharam: se este envio renomeou alguma linha nessa
    # aba, a identidade que falta pode ser a que acabou de nascer — a alteração
    # continua pendente e o envio seguinte, com a folha relida, apanha-a
    for falta in nao_encontradas:
        if falta["sheet"] in renomeadas:
            failed.append({"fn": falta["fn"],
                           "error": "linha não encontrada na folha lida antes "
                                    "deste envio (outra linha foi renomeada "
                                    "agora) - envia outra vez"})
        else:
            failed.append({"fn": falta["fn"], "error": "linha não encontrada na folha"})
    save_overrides(overrides)
    if pushed:
        # o valor local (✎) que segurava a célula acabou de ser apagado: a
        # leitura seguinte tem de vir do livro, senão mostrava as linhas de
        # antes da escrita e o estado parecia ter voltado atrás. Não se confia
        # na data do ficheiro para dar por isso — o Excel pode ainda não a ter
        # atualizado quando o pedido a seguir chegar.
        forget_cache(target)
        # as outras janelas (e os outros dispositivos) estão a mostrar as linhas
        # de antes desta escrita: o aviso põe-nas a reler sem esperar pelo ciclo
        events.publish("sheet", file=os.path.basename(target), pushed=pushed,
                       failed=len(failed))
    return target, pushed, failed


def discard_overrides(target):
    """Descarta as alterações locais (✎) de UM livro e devolve quantas ficaram
    (dos outros livros abertos).

    O critério é o do push_overrides: as chaves deste livro e as antigas (sem
    livro), que são as que um Push a este livro levaria.
    """
    overrides = load_overrides()
    alvo = os.path.normcase(str(target))
    ficam = {}
    for key, entry in overrides.items():
        wb_id, _, _, _ = _split_cellcat_key(key)
        if wb_id is None:
            wb_id, _, _, _ = _split_key(key)
        if wb_id is None or os.path.normcase(str(wb_id)) == alvo:
            continue
        ficam[key] = entry
    save_overrides(ficam)
    forget_cache(target)
    # conta-se o que FICOU, não se relê o ficheiro: é a mesma unidade do
    # pending_all (uma entrada por campo, ver _summarize_overrides)
    return len(_summarize_overrides(ficam))


def pending_overrides_summary():
    """Lista legível de cada alteração local (✎) por enviar — uma entrada por
    campo, não por linha — para o cliente poder mostrar o que vai mesmo ser
    enviado no próximo Push, em vez de só um número.
    Duas formas de chave partilham o mesmo `overrides.json` (ver
    _cellcat_key/_wb_key): a de categoria livre (vista mapeada) guarda um único
    {value,base[,list]} por célula, já as colunas fixas do tracker guardam um
    dict por coluna alterada nessa linha — por isso não se pode confiar em
    len(entry) sem saber qual é.

    Cada entrada leva o LIVRO a que pertence: a lista é de todos os livros, mas
    um Push só escreve num (ver push_overrides), e sem o livro o painel dizia
    "aba · tarefa · campo" de alterações que aquele Push não ia levar.
    O livro é None nas chaves antigas (de quando só havia um livro): essas vão
    para o livro que estiver a ser enviado, seja qual for.
    """
    return _summarize_overrides(load_overrides())


def _summarize_overrides(overrides):
    """O trabalho do pending_overrides_summary sobre um dicionário qualquer, para
    o discard_overrides poder contar o que ficou sem reler o ficheiro."""
    out = []
    for key, entry in overrides.items():
        if not isinstance(entry, dict) or not entry:
            continue
        wb_id, sheet, xlrow, col0 = _split_cellcat_key(key)
        if wb_id is not None:
            headers = known_headers(wb_id, sheet)
            col_i = int(col0)
            field = headers[col_i] if headers and 0 <= col_i < len(headers) else f"Coluna {col_i + 1}"
            out.append({"book": wb_id, "sheet": sheet, "task": f"Linha {xlrow}",
                        "field": field, "value": entry.get("value", ""),
                        "base": entry.get("base", "")})
            continue
        wb_id, sheet, fn, todo = _split_key(key)
        task = fn if not todo or todo == fn else f"{fn} — {todo}"
        for col, sub in entry.items():
            if isinstance(sub, dict):
                out.append({"book": wb_id or "", "sheet": sheet, "task": task,
                            "field": col, "value": sub.get("value", ""),
                            "base": sub.get("base", "")})
    return out


def pending_for_book(details, target):
    """Quantas das alterações locais é que um Push a `target` vai mesmo levar.

    A conta tem de ser a mesma que o push_overrides faz, senão o botão prometia
    um número e escrevia outro: as chaves de outro livro ficam para trás e as
    antigas (sem livro) vão sempre.
    """
    if not target:
        return 0
    alvo = os.path.normcase(target)
    return sum(1 for d in details
               if not d.get("book") or os.path.normcase(str(d["book"])) == alvo)


def _with_app_state(result, target=None):
    """Junta a qualquer resposta o estado que não depende do Excel (CCRs, TODO,
    pendentes, versão da app) — para essas vistas funcionarem à mesma quando
    não há nenhum ficheiro Excel disponível (ou a nuvem pede login).

    `target` é o livro desta resposta (None quando não há nenhum): é o que
    permite dizer quantas das alterações locais este Push vai mesmo levar."""
    result["ccrs"] = load_ccrs()
    # TODO é totalmente manual: colunas/estado só mudam por ação explícita
    # do utilizador (drag/drop, checkbox, botões de cronómetro).
    result["todo"] = load_todo()
    # A lista é de TODOS os livros (o painel mostra-a toda, dizendo de que livro
    # é cada alteração); o número do botão é só o que ESTE Push vai levar — ver
    # pending_for_book e push_overrides.
    detalhes = pending_overrides_summary()
    result["pending_details"] = detalhes
    result["pending_all"] = len(detalhes)
    result["pending"] = pending_for_book(detalhes, target)
    ip = lan_ip()
    if ip:
        result["lan_url"] = f"http://{ip}:{config.SERVER_PORT}"
    result["app_version"] = f"v{APP_VERSION}"
    result["mode"] = "dev" if config.DEV_MODE else "stable"
    return result


def build_payload(query):
    person = query.get("person", [DEFAULT_PERSON])[0]
    sheet = query.get("sheet", [DEFAULT_SHEET])[0]
    show_all = query.get("all", ["0"])[0] == "1"
    wanted_file = query.get("file", [""])[0]
    # cycle=1 (só no botão "Atualizar"): se o Excel local bloquear o
    # ficheiro, fecha-o (gravando), lê os dados frescos e reabre-o
    cycle = query.get("cycle", ["0"])[0] == "1"
    # fresh=1 (botão "Atualizar"): deita fora tudo o que está em memória e lê o
    # livro de raiz, exatamente como na primeira abertura
    fresh = query.get("fresh", ["0"])[0] == "1"
    lang = query.get("lang", ["pt"])[0]
    # fonte dos dados: auto (web primeiro, ficheiro local em recurso),
    # onedrive (só a API do Excel) ou local (só o ficheiro)
    source = query.get("source", ["auto"])[0]
    if source not in ("auto", "onedrive", "local"):
        source = "auto"

    # o livro pedido por esta chamada: se `file` já é um caminho da nuvem (uma
    # aba concreta, aberta explicitamente), é esse — nunca o "atual" singular,
    # senão duas abas OneDrive diferentes abertas ao mesmo tempo mostrariam
    # sempre os dados da última escolhida em qualquer uma delas
    if wanted_file and is_graph_path(wanted_file):
        _wanted_drive_id, _wanted_item_id = graph_ids_from_path(wanted_file)
        # o nome é opcional (só serve para achar a cópia sincronizada local —
        # local_twin/known_files); quem abriu esta aba já o conhecia, do pick
        book = {"drive_id": _wanted_drive_id, "item_id": _wanted_item_id,
                "name": query.get("book_name", [""])[0]} \
            if _wanted_drive_id and _wanted_item_id else current_book()
    else:
        book = current_book()
    files = known_files(book)
    files_info = [{
        "path": p,
        "label": f"{os.path.basename(p)} ({os.path.basename(os.path.dirname(p))})",
        "modified": datetime.fromtimestamp(os.path.getmtime(p)).strftime("%d/%m/%Y %H:%M"),
    } for p in files]

    graph = graph_state()
    # o servidor está exposto na LAN: quem é a conta ligada nunca sai por aqui
    # (o /api/graph só o diz a quem pede deste PC — ver graph_state_public)
    graph_public = graph_state_public(graph)
    # "caminho" do livro na nuvem: o do próprio livro escolhido (identidade que
    # não se confunde com a de outro livro) ou, quando o livro só está indicado
    # na configuração, o caminho geral da fonte web
    web_path = graph_path_for(book["drive_id"], book["item_id"]) if book else GRAPH_PATH
    if graph["configured"]:
        files_info.insert(0, {"path": web_path,
                              "label": graph.get("book") or "OneDrive (web)",
                              "modified": ""})

    path = None
    web_ready = graph["configured"] and graph["connected"] and has_book(book)
    # em automático prefere-se a nuvem (Graph): a cópia sincronizada no disco
    # (twin) só entra como recurso, se a leitura da nuvem falhar mais abaixo —
    # ver o "a fonte web falhou" perto do fim desta função.
    twin = local_twin(files, book) if web_ready else None
    if source != "local" and web_ready:
        path = web_path
    elif source == "onedrive" and graph["configured"]:
        return _with_app_state({"error": msg("err_graph_login", lang),
                "hint": msg("hint_graph_login", lang),
                "files": files_info, "graph": graph_public, "source": "onedrive"})
    if path is None and wanted_file and not is_graph_path(wanted_file):
        for p in files:
            if os.path.normcase(p) == os.path.normcase(wanted_file):
                path = p
                break
    if path is None:
        # nenhum livro pedido (nem nuvem ligada): não é um erro, é o estado
        # normal de quem ainda não escolheu nenhum — a app nunca escolhe por si
        return _with_app_state({
            "no_workbook": True,
            "error": msg("err_nofile", lang),
            "hint": msg("hint_nofile", lang),
            "searched": CANDIDATE_DIRS,
            "files": files_info,
            "graph": graph_public,
        })
    cycle = cycle and not is_graph_path(path)   # não há Excel local para fechar
    if fresh:
        # "Atualizar" limpa tudo o que está em memória, não só do ficheiro
        # atual: qualquer outro livro aberto nesta sessão também relê de raiz
        forget_cache(None)
        graph_forget_item()
        log_event(f"leitura de raiz pedida pelo utilizador ({os.path.basename(path)})")
    cache_key = (path, sheet, person, show_all, lang)
    sheet_read = sheet    # aba pedida à read_sheet (é a chave do _RAW_CACHE)
    try:
        result = read_sheet(path, sheet, person, show_all, lang)
        if result.get("error") and result.get("sheets"):
            # livro diferente do habitual: abre a primeira aba em vez de
            # falhar por a aba pedida não existir
            primeira = result["sheets"][0]
            sheet_read = primeira
            result = read_sheet(path, primeira, person, show_all, lang)
            if "error" not in result:
                result["notice"] = msg("notice_sheet", lang, s=primeira)
        if cycle and result.get("warning"):
            # o utilizador pediu "Atualizar" com o ficheiro bloqueado:
            # fechar o Excel (gravando), reler e reabrir
            log_event("ficheiro bloqueado no Atualizar - a fechar o Excel para atualizar")
            if close_excel_workbook(os.path.basename(path)):
                time.sleep(1.0)
                try:
                    result = read_sheet(path, sheet, person, show_all, lang)
                    result["notice"] = msg("notice_cycled", lang)
                finally:
                    os.startfile(path)
                log_event("Excel fechado, dados atualizados, Excel reaberto")
            else:
                log_event("não consegui fechar o Excel via COM")
        if "error" not in result and not result.get("warning"):
            _LAST_GOOD[cache_key] = (datetime.now(), result)
    except Exception as exc:  # ficheiro bloqueado pelo Excel, corrompido, etc.
        result = None
        if cycle:
            log_event("ficheiro bloqueado no Atualizar - a fechar o Excel para atualizar")
            if close_excel_workbook(os.path.basename(path)):
                time.sleep(1.0)
                try:
                    result = read_sheet(path, sheet, person, show_all, lang)
                    if "error" not in result:
                        _LAST_GOOD[cache_key] = (datetime.now(), result)
                        result["notice"] = msg("notice_cycled", lang)
                    os.startfile(path)
                    log_event("Excel fechado, dados atualizados, Excel reaberto")
                except Exception as exc2:
                    exc = exc2
                    result = None
                    log_event(f"leitura falhou mesmo depois de fechar o Excel ({exc2!r})")
            else:
                log_event("não consegui fechar o Excel via COM")
        if result is not None:
            pass  # leitura fresca conseguida com o ciclo fechar/reabrir
        elif (cached := _LAST_GOOD.get(cache_key)):
            ts, good = cached
            result = dict(good)
            result["warning"] = msg(
                "warning_web" if is_graph_path(path) else "warning_locked",
                lang, t=f"{ts:%H:%M}")
            log_event(f"leitura falhou ({exc!r}) - a servir cache das {ts:%H:%M}")
        elif (retrato := load_last_read(cache_key)):
            # a app acabou de arrancar e nunca conseguiu ler: o retrato da última
            # leitura que correu bem é melhor do que uma vista vazia, que parece
            # "não tens nada" em vez de "não consegui ler" (ver save_last_read)
            quando = str(retrato.get("at") or "")
            result = {k: v for k, v in retrato.items() if k != "at"}
            result["snapshot"] = quando
            result["warning"] = msg("warning_snapshot", lang,
                                    t=quando.replace("T", " ")[:16])
            log_event(f"leitura falhou ({exc!r}) - a servir o retrato de {quando}")
        else:
            result = {"error": msg("err_read", lang, e=exc),
                      "hint": msg("hint_web_read" if is_graph_path(path) else "hint_excel", lang)}

    result["file"] = path
    if is_graph_path(path):
        if "error" in result and source == "auto" and twin:
            # a fonte web falhou: continua com a cópia sincronizada do livro
            # escolhido (o único ficheiro local que se sabe ser o mesmo livro)
            log_event(f"leitura web falhou ({result['error']}) - a usar o ficheiro local")
            fallback = dict(query)
            fallback["source"] = ["local"]
            fallback["file"] = [twin]
            result = build_payload(fallback)
            result["notice"] = msg("notice_graph_fallback", lang)
            result["graph"] = graph_public
            return result
        try:
            result["modified"], result["stamp"] = graph_modified(*graph_ids_from_path(path))
        except GraphError:
            result["modified"] = result["stamp"] = ""
    else:
        mtime = os.path.getmtime(path)
        result["modified"] = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
        result["stamp"] = f"{mtime:.0f}"
        # a ler a cópia sincronizada: avisa quando a do OneDrive ainda está
        # diferente, para ninguém ficar a pensar que os dados estão parados
        cached = _RAW_CACHE.get((path, normalize(sheet_read)))
        if twin and path == twin and cached and not result.get("error"):
            try:
                if sync_gap(path, result["sheet"], cached[3], mtime,
                            *graph_ids_from_path(web_path)):
                    aviso = msg("notice_syncing", lang)
                    result["notice"] = f"{result['notice']} · {aviso}" \
                        if result.get("notice") else aviso
            except GraphError as exc:
                log_event(f"não consegui comparar com a cópia do OneDrive ({exc})")
    result["files"] = files_info
    result["graph"] = graph_public
    result["source"] = "onedrive" if is_graph_path(path) else "local"
    result["synced_copy"] = bool(twin) and path == twin
    result = _with_app_state(result, path)
    # impressão digital do conteúdo servido: se o Excel mudou e isto não muda,
    # o problema está na origem (livro por gravar), não na app
    result["digest"] = hashlib.md5(
        json.dumps(result.get("rows") or [], ensure_ascii=False).encode("utf-8")
    ).hexdigest()[:8]

    # vista resumida à medida por coordenadas de célula (ver viewmap.js): o
    # cliente manda a lista de categorias guardada para este ficheiro+aba: só
    # faz sentido calcular com uma leitura sem erro (row_meta vem de lá)
    if not result.get("error"):
        cellcats_raw = query.get("cellcats", [""])[0]
        if cellcats_raw:
            try:
                categories = json.loads(cellcats_raw)
            except (TypeError, ValueError):
                categories = None
            if isinstance(categories, list) and categories:
                result["cell_view"] = build_cell_categories(
                    path, sheet_read, categories, result.get("row_meta"))

        # filtros personalizados (ver customfilters.js): listas predefinidas
        # mode="range" referenciadas por um filtro in_list/not_in_list — só o
        # servidor consegue ler o intervalo ao vivo, tal como para as
        # categorias listMode="fixed" acima
        filterlists_raw = query.get("filterlists", [""])[0]
        if filterlists_raw:
            try:
                wanted_lists = json.loads(filterlists_raw)
            except (TypeError, ValueError):
                wanted_lists = None
            if isinstance(wanted_lists, list) and wanted_lists:
                resolved = {}
                for entry in wanted_lists:
                    if not isinstance(entry, dict):
                        continue
                    list_id = str(entry.get("id") or "").strip()
                    list_sheet = str(entry.get("sheet") or "").strip()
                    list_cell = str(entry.get("cell") or "").strip()
                    if not list_id or not list_sheet or not list_cell:
                        continue
                    resolved[list_id] = _read_list_options(
                        path, list_sheet, list_cell, entry.get("orientation"), entry.get("size"))
                if resolved:
                    result["filter_lists"] = resolved

    # o retrato para o próximo arranque às escuras (ver save_last_read): grava-se
    # no fim, com a leitura já completa, e só quando o conteúdo mudou
    if not result.get("error") and not result.get("snapshot"):
        try:
            save_last_read(cache_key, result)
        except OSError as exc:
            log_event(f"não consegui guardar o retrato da leitura ({exc})")

    return result


def current_stamp(query):
    """Marca de versão da fonte em uso, sem ler a folha. É o pedido barato que
    a interface repete para saber quando alguém gravou o livro."""
    path = query.get("file", [""])[0]
    try:
        if is_graph_path(path):
            modified, stamp = graph_modified(*graph_ids_from_path(path))
            return {"modified": modified, "stamp": stamp}
        # só ficheiros que a app conhece: evita transformar isto num "stat" livre
        if any(os.path.normcase(p) == os.path.normcase(path) for p in known_files()):
            mtime = os.path.getmtime(path)
            return {"modified": datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M"),
                    "stamp": f"{mtime:.0f}"}
    except Exception as exc:
        return {"modified": "", "stamp": "", "error": str(exc)}
    return {"modified": "", "stamp": ""}


# ---------------------------------------------------------------------------
# O que o SERVIDOR já leu, para o assistente poder perguntar mais
#
# O motor LLM recebia no pedido as primeiras 120 linhas do retrato que o cliente
# mandou e não tinha maneira de pedir mais: numa folha grande respondia sobre
# essa parte, sem saber que havia outra. Isto dá-lhe uma janela para o que o
# servidor tem em memória — TODAS as linhas de TODAS as folhas já lidas, de
# todos os livros abertos, e não só as do separador de quem está a perguntar.
#
# Não lê nada: só o _RAW_CACHE, que é o resultado da última leitura de cada
# folha. A regra de o assistente nunca ir à folha (nem ao disco, nem à nuvem)
# para responder a uma pergunta fica de pé.


def cached_books():
    """As folhas que o servidor tem em memória: (livro, aba, linhas, quando)."""
    out = []
    for (path, _chave), entry in list(_RAW_CACHE.items()):
        try:
            quando, real_sheet, _all_sheets, rows = entry
        except (TypeError, ValueError):
            continue
        out.append({"book": os.path.basename(path), "sheet": real_sheet or "",
                    "rows": max(0, len(rows or []) - 1),
                    "read": quando.strftime("%Y-%m-%d %H:%M") if quando else ""})
    out.sort(key=lambda b: (b["book"], b["sheet"]))
    return out


def cached_rows(book="", sheet="", query="", limit=20, offset=0, col_chars=80):
    """Linhas das folhas em memória, em texto. Devolve {rows, total, books}.

    `query` são palavras que têm todas de aparecer na linha (sem acentos, como
    o resto da app procura). `total` é quantas correspondem — o assistente sabe
    assim que há mais para pedir com o `offset`, em vez de achar que viu tudo.
    """
    termos = [normalize(t) for t in str(query or "").split() if t]
    quer_livro, quer_aba = normalize(book or ""), normalize(sheet or "")
    limite = max(1, min(60, int(limit or 20)))
    salto = max(0, int(offset or 0))
    achadas, total = [], 0
    for (path, _chave), entry in sorted(_RAW_CACHE.items(), key=lambda kv: str(kv[0])):
        try:
            _quando, real_sheet, _all_sheets, rows = entry
        except (TypeError, ValueError):
            continue
        nome = os.path.basename(path)
        if quer_livro and quer_livro not in normalize(nome):
            continue
        if quer_aba and quer_aba not in normalize(real_sheet or ""):
            continue
        idx = detect_header_row(rows or [])
        if idx is None:
            continue
        cabecalhos = [cell_to_text(c) for c in rows[idx]]
        for n, raw in enumerate(rows[idx + 1:], start=idx + 2):
            partes = []
            for cab, valor in zip(cabecalhos, list(raw or [])):
                texto = cell_to_text(valor)
                if cab and texto:
                    partes.append(f"{cab}: {texto[:col_chars]}")
            if not partes:
                continue
            linha = " | ".join(partes)[:600]
            if termos and not all(t in normalize(linha) for t in termos):
                continue
            total += 1
            if total <= salto or len(achadas) >= limite:
                continue
            achadas.append({"book": nome, "sheet": real_sheet or "",
                            "xlrow": n, "text": linha})
    return {"rows": achadas, "total": total, "books": cached_books()}
