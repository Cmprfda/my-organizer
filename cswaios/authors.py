# -*- coding: utf-8 -*-
"""Quem mudou ESTA célula.

O histórico (cswaios/history.py) sabe o que mudou, quando, e se foi esta app a
escrever. O que não sabia era de quem foi a alteração que veio de fora: o Excel
não guarda autoria por célula e o OneDrive só sabe quem gravou o ficheiro. Até
aqui a app cruzava as horas — o `☁` de uma alteração das 14:03 mostrava o nome
de quem gravou às 14:03 — e duas pessoas a gravar no mesmo minuto (ou uma a
gravar por cima do trabalho da outra, em coautoria) davam o mesmo nome a
alterações que não eram dela.

Aqui vai-se ver. Cada versão do livro no OneDrive é um ficheiro inteiro: pega-se
na versão candidata, abre-se, e lê-se A CÉLULA. Se ela já tem o valor novo e a
versão anterior ainda tem o antigo, a alteração nasceu naquela gravação e o
autor é o dela — sem adivinhar por horas.

Custa uma descarga do livro por versão consultada (são megabytes), por isso:
  - só a pedido, quando alguém clica no `☁` de uma alteração;
  - só as versões perto da hora da alteração;
  - com o resultado e os bytes guardados em memória, que a pergunta seguinte
    sobre o mesmo Push é sobre as mesmas versões.
"""

import io
import threading
import time
import urllib.error
import urllib.parse
import urllib.request

import openpyxl

from .excel import detect_header_row
from .graph import (GraphError, graph_config, graph_ids_from_path, graph_token,
                    graph_versions, is_graph_path)
from .text import cell_to_text, normalize

# versões a consultar por pergunta: a alteração está entre a gravação anterior à
# hora dela e as que vieram logo depois. Mais do que isto era descarregar o
# livro meia dúzia de vezes para responder a uma pergunta.
MAX_VERSOES = 4

# margem em torno da hora da alteração. O histórico marca a alteração quando a
# LEITURA a encontrou, que pode ser bem depois de ela ser gravada (coautoria do
# OneDrive, ver docs) — daí a janela ser larga para trás.
MARGEM_ANTES = 6 * 3600
MARGEM_DEPOIS = 30 * 60

# bytes de versões já descarregadas: (drive, item, versão) -> (quando, bytes).
# Duas ou três, e as mais antigas saem — um livro destes são megabytes.
_BYTES_CACHE = {}
_BYTES_MAX = 3

# respostas já dadas: a chave é a alteração (livro, aba, linha, coluna, hora)
_QUEM_CACHE = {}
_QUEM_TTL = 12 * 3600

_lock = threading.Lock()


class AuthorError(Exception):
    """Não se consegue responder (ficheiro local, sem sessão, sem versões)."""


def _iso_ts(texto):
    """Segundos desde a época a partir de uma marca ISO (0 se não presta)."""
    texto = str(texto or "").strip()
    if not texto:
        return 0
    if texto.endswith("Z"):
        texto = texto[:-1] + "+00:00"
    try:
        from datetime import datetime
        quando = datetime.fromisoformat(texto)
        if quando.tzinfo is None:
            return quando.timestamp()          # marca local (a do histórico)
        return quando.timestamp()
    except ValueError:
        return 0


def _version_bytes(drive_id, item_id, version_id):
    """O livro tal como estava naquela versão."""
    chave = (drive_id, item_id, version_id)
    with _lock:
        em_cache = _BYTES_CACHE.get(chave)
        if em_cache:
            return em_cache[1]
    cfg = graph_config()
    if not cfg:
        raise AuthorError("acesso web não configurado")
    token = graph_token(cfg)
    if not token:
        raise AuthorError("sessão do OneDrive não iniciada")
    url = (f"{cfg['graph_base']}/drives/{drive_id}/items/{item_id}"
           f"/versions/{urllib.parse.quote(str(version_id))}/content")
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = resp.read()
    except (urllib.error.HTTPError, urllib.error.URLError, OSError) as exc:
        raise AuthorError(f"não deu para ler a versão: {exc}") from exc
    with _lock:
        _BYTES_CACHE[chave] = (time.time(), data)
        if len(_BYTES_CACHE) > _BYTES_MAX:
            velha = min(_BYTES_CACHE, key=lambda k: _BYTES_CACHE[k][0])
            _BYTES_CACHE.pop(velha, None)
    return data


def _cell_in_bytes(data, sheet, xlrow, col_name):
    """O valor daquela célula naquele livro. None quando não se encontra.

    A coluna é procurada pelo NOME (é o que o histórico guarda) na linha de
    cabeçalho daquela versão: entre duas versões, uma coluna pode ter mudado de
    lugar, e a mesma letra deixaria de querer dizer a mesma coisa.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:                      # ficheiro cortado a meio
        raise AuthorError(f"versão ilegível: {exc}") from exc
    try:
        alvo = normalize(sheet)
        nome = next((n for n in wb.sheetnames if normalize(n) == alvo), None) \
            or next((n for n in wb.sheetnames if alvo in normalize(n)), None)
        if nome is None:
            return None
        # read_only: anda-se uma vez pelas linhas até à que interessa
        linhas = []
        for n, row in enumerate(wb[nome].iter_rows(values_only=True), start=1):
            linhas.append(list(row))
            if n >= int(xlrow):
                break
        if len(linhas) < int(xlrow):
            return None
        cabecalho = detect_header_row(linhas)
        if cabecalho is None:
            return None
        quer = normalize(col_name)
        col = next((i for i, c in enumerate(linhas[cabecalho])
                    if normalize(cell_to_text(c)) == quer), None)
        if col is None:
            return None
        linha = linhas[int(xlrow) - 1]
        return cell_to_text(linha[col]) if col < len(linha) else ""
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _candidatas(versoes, quando):
    """As versões que podem ter trazido uma alteração vista às `quando`.

    Das mais antigas para as mais novas, porque é essa a ordem em que a
    alteração aparece: a primeira que já tem o valor novo é a que o trouxe.
    """
    alvo = _iso_ts(quando)
    if not alvo:
        return list(reversed(versoes))[-MAX_VERSOES:]
    perto = [v for v in versoes
             if -MARGEM_ANTES <= _iso_ts(v.get("when")) - alvo <= MARGEM_DEPOIS
             or _iso_ts(v.get("when")) <= alvo]
    perto.sort(key=lambda v: _iso_ts(v.get("when")))
    return perto[-MAX_VERSOES:]


def who_changed(livro, sheet, xlrow, col, ts, antes="", depois=""):
    """Quem gravou a versão que trouxe esta alteração.

    Devolve {who, when, version, confirmed, checked} — `confirmed` diz que a
    versão anterior ainda tinha o valor antigo, ou seja, que a alteração nasceu
    ali mesmo e não numa gravação de que já não há versão.
    """
    if not is_graph_path(livro):
        raise AuthorError("só a fonte web guarda versões do livro")
    chave = (normalize(livro), normalize(sheet), int(xlrow), str(col),
             str(ts), str(depois))
    agora = time.time()
    with _lock:
        em_cache = _QUEM_CACHE.get(chave)
        if em_cache and agora - em_cache[0] < _QUEM_TTL:
            return em_cache[1]
    drive_id, item_id = graph_ids_from_path(livro)
    if not (drive_id and item_id):
        raise AuthorError("livro desconhecido")
    try:
        versoes = graph_versions(drive_id, item_id)
    except GraphError as exc:
        raise AuthorError(str(exc)) from exc
    versoes = [v for v in versoes if v.get("id")]
    if not versoes:
        raise AuthorError("o OneDrive não deu versões deste livro")
    candidatas = _candidatas(versoes, ts)
    if not candidatas:
        raise AuthorError("nenhuma gravação perto da hora da alteração")
    esperado = str(depois or "")
    vistas = []
    anterior_valor = None
    for v in candidatas:
        try:
            valor = _cell_in_bytes(_version_bytes(drive_id, item_id, v["id"]),
                                   sheet, xlrow, col)
        except AuthorError:
            continue                    # uma versão que não se lê não invalida as outras
        vistas.append({"when": v.get("when"), "who": v.get("who"), "value": valor})
        if valor is not None and esperado and valor.strip() == esperado.strip():
            out = {
                "who": v.get("who") or "",
                "when": v.get("when") or "",
                "version": str(v["id"]),
                # a versão de antes ainda tinha o valor antigo: a alteração
                # nasceu MESMO nesta gravação
                "confirmed": anterior_valor is not None
                and anterior_valor.strip() == str(antes or "").strip(),
                "checked": len(vistas),
            }
            with _lock:
                _QUEM_CACHE[chave] = (agora, out)
            return out
        anterior_valor = valor
    raise AuthorError("nenhuma das gravações consultadas tem este valor")
