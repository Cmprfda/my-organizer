# -*- coding: utf-8 -*-
"""Leitura do .xlsx (openpyxl) e escrita através do Excel/COM."""

import glob
import json
import os
import re
import subprocess
import tempfile
import time
from datetime import datetime

import openpyxl

from . import config
from .config import CANDIDATE_DIRS, FILE_PATTERN, HERE
from .graph import (graph_ids_from_path, graph_load_rows, graph_write_status,
                    is_graph_path)
from .text import cell_to_text, normalize

# folha crua por (ficheiro, aba) — independente da pessoa/vista, para que
# qualquer dispositivo seja servido a partir da última leitura do ficheiro
_RAW_CACHE = {}


# lista oficial de estados, lida da coluna "Status" da aba Admin do próprio
# ficheiro (cache por mtime, para não reabrir o workbook a cada pedido)
_ADMIN_CACHE = {}


def admin_statuses(path):
    if is_graph_path(path):
        mtime = int(time.time() // 300)   # fonte web: revalida a cada 5 minutos
    else:
        try:
            mtime = os.path.getmtime(path)
        except OSError:
            mtime = None
    cached = _ADMIN_CACHE.get(path)
    if cached and cached[0] == mtime:
        return cached[1]
    try:
        if is_graph_path(path):
            drive_id, item_id = graph_ids_from_path(path)
            sheet, _all, rows = graph_load_rows(drive_id, item_id, "Admin")
            if sheet is None:
                return cached[1] if cached else None
        else:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = pick_sheet(wb, "Admin")
                if sheet is None:
                    return cached[1] if cached else None
                rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
            finally:
                wb.close()
    except Exception:  # bloqueado pelo Excel / sem rede — usa a última leitura boa
        return cached[1] if cached else None

    col, start = None, 0
    for i, row in enumerate(rows[:10]):
        for j, v in enumerate(row):
            if isinstance(v, str) and normalize(v) == "status":
                col, start = j, i + 1
                break
        if col is not None:
            break
    if col is None:
        return cached[1] if cached else None
    vals = []
    for row in rows[start:]:
        v = cell_to_text(row[col]) if col < len(row) else ""
        if not v:
            break  # a lista termina na primeira célula vazia
        if v not in vals:
            vals.append(v)
    if vals:
        _ADMIN_CACHE[path] = (mtime, vals)
        return vals
    return cached[1] if cached else None



EXCEL_WRITE_PS1 = r"""
param([string]$ParamsPath)
$ErrorActionPreference = 'Stop'
$own = $null
try {
  $p = Get-Content -Raw -Path $ParamsPath -Encoding UTF8 | ConvertFrom-Json
  $wb = $null
  try { $x = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application') } catch { $x = $null }
  if ($x) {
    $wb = @($x.Workbooks) | Where-Object { $_.Name -eq $p.basename } | Select-Object -First 1
  }
  if (-not $wb) {
    $own = New-Object -ComObject Excel.Application
    $own.Visible = $false
    $own.DisplayAlerts = $false
    $wb = $own.Workbooks.Open($p.path)
  }
  $ws = $wb.Worksheets.Item($p.sheet)
  $fnCell = [string]$ws.Cells($p.xlrow, $p.fncol).Value2
  $a = ($fnCell -replace '\s+', ' ').Trim()
  $b = ([string]$p.fn -replace '\s+', ' ').Trim()
  if ($a -ne $b) {
    throw "a linha $($p.xlrow) da folha mudou entretanto (esperava '$b', encontrei '$a') - atualiza a app e tenta de novo"
  }
  $cell = $ws.Cells($p.xlrow, $p.xlcol)
  $cell.Value2 = [string]$p.value
  # texto com mudanças de linha só se lê na folha com "moldar texto" ligado:
  # sem isto o Excel guardava as linhas mas mostrava-as todas colada numa só
  if (([string]$p.value).Contains([string][char]10)) { $cell.WrapText = $true }
  $wb.Save()
  if ($own) { $wb.Close($true); $own.Quit() }
  Write-Output 'OK'
  exit 0
} catch {
  Write-Output ('ERRO: ' + $_.Exception.Message)
  if ($own) { try { $wb.Close($false) } catch {}; try { $own.Quit() } catch {} }
  exit 1
}
"""


def write_status_to_excel(path, sheet, xlrow, xlcol, fncol, fn, value):
    """Escreve um estado diretamente na folha, através do Excel (COM), que
    preserva gráficos/validações e faz upload via OneDrive. Usa o livro já
    aberto se existir; senão abre uma instância invisível só para isto.
    Devolve (ok, mensagem)."""
    # dentro de uma célula o Excel muda de linha com \n; um \r a acompanhar
    # aparece na folha como um quadradinho no meio do texto
    if isinstance(value, str):
        value = value.replace("\r\n", "\n").replace("\r", "\n")
    if is_graph_path(path):
        # fonte web: o Excel/COM não se aplica, escreve-se pela API do Excel
        drive_id, item_id = graph_ids_from_path(path)
        return graph_write_status(sheet, xlrow, xlcol, fncol, fn, value,
                                  drive_id, item_id)
    params = {"path": path, "basename": os.path.basename(path), "sheet": sheet,
              "xlrow": int(xlrow), "xlcol": int(xlcol), "fncol": int(fncol),
              "fn": fn, "value": value}
    with tempfile.TemporaryDirectory() as td:
        params_path = os.path.join(td, "params.json")
        ps1_path = os.path.join(td, "write.ps1")
        with open(params_path, "w", encoding="utf-8") as f:
            json.dump(params, f, ensure_ascii=False)
        with open(ps1_path, "w", encoding="ascii") as f:
            f.write(EXCEL_WRITE_PS1)
        try:
            proc = subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
                 "-File", ps1_path, params_path],
                capture_output=True, timeout=120)
            out = proc.stdout.decode("utf-8", errors="replace").strip()
            return proc.returncode == 0, out or "sem resposta do Excel"
        except subprocess.TimeoutExpired:
            return False, "o Excel demorou demasiado a responder"
        except Exception as exc:
            return False, str(exc)


def _parse_cell_ref(cell):
    m = re.match(r"^\$?([A-Za-z]+)\$?(\d+)$", (cell or "").strip())
    if not m:
        raise ValueError(f"referência de célula inválida: {cell!r}")
    return m.group(1).upper(), int(m.group(2))


def _col_letters_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _num_to_col_letters(n):
    letters = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        letters = chr(65 + r) + letters
    return letters


def _options_list_range(sheet, cell, orientation, size):
    """Endereço absoluto do intervalo que contém a lista de opções, a partir
    da sua primeira célula, orientação (vertical/horizontal) e nº de itens."""
    col, row = _parse_cell_ref(cell)
    try:
        size = int(size)
    except (TypeError, ValueError):
        raise ValueError("o tamanho da lista tem de ser um número")
    if size < 1:
        raise ValueError("o tamanho da lista tem de ser pelo menos 1")
    if orientation == "vertical":
        end_col, end_row = col, row + size - 1
    elif orientation == "horizontal":
        end_col, end_row = _num_to_col_letters(_col_letters_to_num(col) + size - 1), row
    else:
        raise ValueError("orientação inválida (usa 'vertical' ou 'horizontal')")
    return f"'{sheet}'!${col}${row}:${end_col}${end_row}"


SET_VALIDATION_PS1 = r"""
param([string]$ParamsPath)
$ErrorActionPreference = 'Stop'
$own = $null
try {
  $p = Get-Content -Raw -Path $ParamsPath -Encoding UTF8 | ConvertFrom-Json
  $wb = $null
  try { $x = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application') } catch { $x = $null }
  if ($x) {
    $wb = @($x.Workbooks) | Where-Object { $_.Name -eq $p.basename } | Select-Object -First 1
  }
  if (-not $wb) {
    $own = New-Object -ComObject Excel.Application
    $own.Visible = $false
    $own.DisplayAlerts = $false
    $wb = $own.Workbooks.Open($p.path)
  }
  $ws = $wb.Worksheets.Item($p.targetSheet)
  $cell = $ws.Range($p.targetCell)
  $cell.Validation.Delete()
  $cell.Validation.Add(3, 1, 1, [string]$p.formula, [string]::Empty)
  $cell.Validation.IgnoreBlank = $true
  $cell.Validation.InCellDropdown = $true
  $wb.Save()
  if ($own) { $wb.Close($true); $own.Quit() }
  Write-Output 'OK'
  exit 0
} catch {
  Write-Output ('ERRO: ' + $_.Exception.Message)
  if ($own) { try { $wb.Close($false) } catch {}; try { $own.Quit() } catch {} }
  exit 1
}
"""


def _run_set_validation(path, target_sheet, target_cell, formula):
    """Corre o SET_VALIDATION_PS1 com o Formula1 já pronto (referência de
    intervalo ou lista literal de valores), partilhado por
    set_data_validation_list e set_data_validation_fixed_list."""
    params = {"path": path, "basename": os.path.basename(path), "targetSheet": target_sheet,
              "targetCell": target_cell, "formula": formula}
    with tempfile.TemporaryDirectory() as td:
        params_path = os.path.join(td, "params.json")
        ps1_path = os.path.join(td, "validation.ps1")
        with open(params_path, "w", encoding="utf-8") as f:
            json.dump(params, f, ensure_ascii=False)
        with open(ps1_path, "w", encoding="ascii") as f:
            f.write(SET_VALIDATION_PS1)
        try:
            proc = subprocess.run(
                ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
                 "-File", ps1_path, params_path],
                capture_output=True, timeout=120)
            out = proc.stdout.decode("utf-8", errors="replace").strip()
            return proc.returncode == 0, out or "sem resposta do Excel"
        except subprocess.TimeoutExpired:
            return False, "o Excel demorou demasiado a responder"
        except Exception as exc:
            return False, str(exc)


def set_data_validation_list(path, target_sheet, target_cell, source_sheet, source_cell,
                              orientation, size):
    """Aplica, a uma célula, uma lista de valores predefinidos (Data
    Validation) cujas opções já existem numa aba/intervalo do próprio livro.
    Escreve pelo Excel (COM), tal como write_status_to_excel, para preservar
    o resto do livro (gráficos, outras validações). Devolve (ok, mensagem)."""
    if is_graph_path(path):
        return False, "não é possível definir validações num ficheiro só acessível via Graph/OneDrive"
    if not target_sheet or not target_cell:
        return False, "aba e célula de destino são obrigatórias"
    if not source_sheet:
        return False, "aba da lista de opções é obrigatória"
    try:
        _parse_cell_ref(target_cell)
        formula = "=" + _options_list_range(source_sheet, source_cell, orientation, size)
    except ValueError as exc:
        return False, str(exc)
    return _run_set_validation(path, target_sheet, target_cell, formula)


def set_data_validation_fixed_list(path, target_sheet, target_cell, values):
    """Aplica, a uma célula, uma lista de valores predefinidos (Data
    Validation) cujas opções são literais (biblioteca de listas guardada no
    cliente, ver viewmap.js), sem depender de nenhum intervalo do livro.
    Devolve (ok, mensagem)."""
    if is_graph_path(path):
        return False, "não é possível definir validações num ficheiro só acessível via Graph/OneDrive"
    if not target_sheet or not target_cell:
        return False, "aba e célula de destino são obrigatórias"
    vals = [str(v).strip() for v in (values or []) if str(v).strip()]
    if not vals:
        return False, "a lista de valores não pode estar vazia"
    if any("," in v for v in vals):
        return False, "os valores da lista não podem conter vírgulas (separador do Excel)"
    try:
        _parse_cell_ref(target_cell)
    except ValueError as exc:
        return False, str(exc)
    formula = ",".join(vals)
    if len(formula) > 255:
        return False, "a lista de valores excede o limite de 255 carateres do Excel"
    return _run_set_validation(path, target_sheet, target_cell, formula)


def locate_row(path, sheet_wanted, fn, todo):
    """Localiza (xlrow, hidx) de uma tarefa na folha atual (ou na cache, se o
    ficheiro estiver bloqueado). Devolve None se a linha não existir."""
    raw_key = (path, normalize(sheet_wanted))
    try:
        if is_graph_path(path):
            drive_id, item_id = graph_ids_from_path(path)
            real_sheet, all_sheets, rows = graph_load_rows(drive_id, item_id, sheet_wanted)
            if real_sheet is None:
                return None
        else:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                real_sheet = pick_sheet(wb, sheet_wanted)
                if real_sheet is None:
                    return None
                rows = [list(r) for r in wb[real_sheet].iter_rows(values_only=True)]
                all_sheets = wb.sheetnames
            finally:
                wb.close()
        _RAW_CACHE[raw_key] = (datetime.now(), real_sheet, all_sheets, rows)
    except Exception:
        if raw_key not in _RAW_CACHE:
            return None
        _, real_sheet, all_sheets, rows = _RAW_CACHE[raw_key]

    header_index = detect_header_row(rows)
    if header_index is None:
        return None
    # mesmo fallback "Coluna N" que tasks.read_sheet/known_headers expõem ao
    # cliente para cabeçalhos vazios — sem isto, colunas sem texto no cabeçalho
    # colidiam todas na mesma chave "" e nunca se encontrava a coluna certa
    headers = [cell_to_text(h) or f"Coluna {i + 1}" for i, h in enumerate(rows[header_index])]
    # todas as colunas reais (normalizadas), não só as 5 do tracker — para a
    # vista mapeada à medida (viewmap.js) poder escrever numa coluna qualquer
    # no Push, identificada pelo nome (ver tasks.push_overrides)
    hidx = {}
    for j, h in enumerate(headers):
        hidx[normalize(h)] = j
    if "function/tc" not in hidx:
        return None
    for i, row in enumerate(rows[header_index + 1:]):
        cells = [cell_to_text(v) for v in row]
        cells += [""] * (len(headers) - len(cells))
        fn_key = cells[hidx["function/tc"]]
        todo_key = cells[hidx["to do"]] if "to do" in hidx else ""
        if fn_key == fn and todo_key == todo:
            return header_index + 2 + i, hidx
    return None


def close_excel_workbook(basename):
    """Pede ao Excel (via COM) para fechar o livro, gravando alterações.
    Devolve True se o livro foi fechado."""
    script = (
        "try { $x = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application') } "
        "catch { exit 2 }; "
        f"$wb = @($x.Workbooks) | Where-Object {{ $_.Name -eq '{basename}' }}; "
        "if (-not $wb) { exit 3 }; "
        "$wb | ForEach-Object { $_.Close($true) }; exit 0"
    )
    try:
        rc = subprocess.run(["powershell", "-NoProfile", "-Command", script],
                            capture_output=True, timeout=40).returncode
        return rc == 0
    except Exception:
        return False


# A procura abaixo varre as pastas candidatas de cima a baixo (as do OneDrive
# podem ser enormes, e com ficheiros só na nuvem cada entrada custa ainda mais).
# É pedida em cada /api/tasks e também em cada /api/modified — o pedido "barato"
# que a interface repete de 20 em 20 segundos por cada livro aberto — por isso
# guarda-se o resultado. Volta a varrer quando alguma pasta candidata mexe (um
# ficheiro novo lá dentro muda a data da pasta: é o caso do download do
# SharePoint, que aterra direito na Downloads) ou passados _FILES_TTL segundos,
# para apanhar o que aterre numa subpasta, onde a data da pasta de cima não
# mexe. As datas mostradas na interface não passam por aqui (o build_payload
# lê o mtime de cada ficheiro a cada pedido); o que pode ficar até _FILES_TTL
# segundos velho é só a LISTA de ficheiros e a ordem entre eles.
_FILES_TTL = 30.0
_FILES_CACHE = {}   # chave -> (assinatura das pastas, momento da leitura, lista)


def _dirs_signature():
    """Data de cada pasta candidata — muda assim que lá aterra um ficheiro."""
    marcas = []
    for base in list(CANDIDATE_DIRS) + [HERE]:
        try:
            marcas.append(os.stat(base).st_mtime_ns)
        except OSError:
            marcas.append(None)
    return tuple(marcas)


def _cached_scan(key, scan):
    """`scan()` guardado enquanto as pastas candidatas não mexerem."""
    agora = time.time()
    sig = _dirs_signature()
    em_cache = _FILES_CACHE.get(key)
    if em_cache and em_cache[0] == sig and agora - em_cache[1] < _FILES_TTL:
        return list(em_cache[2])
    achados = scan()
    _FILES_CACHE[key] = (sig, agora, list(achados))
    return achados


def forget_files_cache():
    """Força a próxima procura a varrer as pastas outra vez."""
    _FILES_CACHE.clear()


def find_tracker_files():
    """Todos os candidatos, do mais recente para o mais antigo."""
    if config.FORCED_FILE:
        return [config.FORCED_FILE] if os.path.isfile(config.FORCED_FILE) else []
    return _cached_scan(("tracker",), _scan_tracker_files)


def _scan_tracker_files():
    matches = set()
    for base in CANDIDATE_DIRS:
        if not os.path.isdir(base):
            continue
        # ~ prefixados são ficheiros temporários de lock do Office
        for path in glob.glob(os.path.join(base, "**", FILE_PATTERN), recursive=True):
            if not os.path.basename(path).startswith("~"):
                matches.add(os.path.abspath(path))
    # qualquer xlsx posto na pasta da app também conta
    for path in glob.glob(os.path.join(HERE, "*.xlsx")):
        if not os.path.basename(path).startswith("~"):
            matches.add(os.path.abspath(path))
    return sorted(matches, key=os.path.getmtime, reverse=True)


def find_named_file(name):
    """Ficheiros com este nome exato nas pastas candidatas, do mais recente
    para o mais antigo. Serve para achar a cópia sincronizada de um livro do
    OneDrive, que pode não ser o tracker (FILE_PATTERN não o apanharia)."""
    if not name:
        return []
    return _cached_scan(("named", name), lambda: _scan_named_file(name))


def _scan_named_file(name):
    pattern = glob.escape(name)
    matches = set()
    for base in CANDIDATE_DIRS:
        if not os.path.isdir(base):
            continue
        for path in glob.glob(os.path.join(base, "**", pattern), recursive=True):
            if os.path.isfile(path) and not os.path.basename(path).startswith("~"):
                matches.add(os.path.abspath(path))
    return sorted(matches, key=os.path.getmtime, reverse=True)


def browse_local_file():
    """Abre a janela do Windows para escolher um .xlsx no disco. Devolve o
    caminho escolhido, None se o utilizador cancelou, ou "unavailable" quando a
    app está a correr no browser (sem janela nativa não há diálogo possível)."""
    try:
        import webview
    except ImportError:
        return "unavailable"
    if not config.WEBVIEW_WINDOW:
        return "unavailable"
    escolhido = config.WEBVIEW_WINDOW.create_file_dialog(
        webview.OPEN_DIALOG, file_types=("Excel files (*.xlsx)",))
    if not escolhido:
        return None
    return str(escolhido[0]) if isinstance(escolhido, (list, tuple)) else str(escolhido)


def pick_sheet(workbook, wanted):
    if wanted in workbook.sheetnames:
        return wanted
    wanted_norm = normalize(wanted)
    for name in workbook.sheetnames:
        if normalize(name) == wanted_norm:
            return name
    for name in workbook.sheetnames:
        if wanted_norm in normalize(name):
            return name
    return None


def detect_header_row(rows):
    """Devolve o índice da linha que parece ser o cabeçalho: a primeira,
    entre as 15 iniciais, com o maior número de células de texto preenchidas."""
    best_index, best_score = 0, -1
    for i, row in enumerate(rows[:15]):
        score = sum(1 for v in row if isinstance(v, str) and v.strip())
        if score > best_score and score >= 2:
            best_index, best_score = i, score
    return best_index if best_score >= 2 else None
